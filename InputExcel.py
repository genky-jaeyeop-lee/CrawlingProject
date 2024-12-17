# InputExcel.py
import openpyxl
import os


def save_to_excel(data, file_path):
    try:
        # ファイル パスがない場合は、新たに作成します
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()  # 新しいエクセルワークブック作成
        else:
            wb = openpyxl.load_workbook(file_path)  # 既存ワークブック開く

        ws = wb.active

        # 最初の行にタイトルを追加（タイトルがなければ追加）
        if ws.max_row == 1:
            ws.append(["コース概要"])

        # クローリングしたデータをエクセルに追加します
        for item in data:
            ws.append([item.get("コース概要", "N/A")])

        # Excelファイル保存
        wb.save(file_path)
        print(f"Excelファイルが正常にアップデートされました: {file_path}")

    except Exception as e:
        print(f"Excelアップデート中にエラーが発生します: {e}")
